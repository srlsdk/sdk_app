"""
core/suivi_mo.py
-----------------
Génère un récapitulatif des jours travaillés par chantier à partir d'un
fichier Excel "Suivi heures".

Logique métier identique au script CLI d'origine (suivi_mo.py), adaptée
pour être appelée depuis Streamlit (entrées/sorties en mémoire).
"""

import io
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SHEET_NAME = "Suivi heures"

SKIP_NAMES = {
    'Noms', 'semaine', 'jour', 'date',
    'ENT Sadaka', 'Suivi des heures',
}


def traiter(input_path: str, mois_label: str = "", log=print) -> bytes:
    """
    input_path : chemin vers le fichier xlsx source (temp dir).
    mois_label : libellé affiché dans le titre (optionnel).
    Retourne le contenu xlsx du récapitulatif en bytes.
    """
    skip_names = SKIP_NAMES | {mois_label} if mois_label else SKIP_NAMES

    wb_src = openpyxl.load_workbook(input_path, data_only=True)
    ws_src = wb_src[SHEET_NAME]

    result = defaultdict(lambda: defaultdict(int))

    for row in ws_src.iter_rows(values_only=True):
        vals = list(row)
        name = vals[1]  # colonne B = nom du salarié
        if not name or str(name).strip() in skip_names:
            continue
        name = str(name).strip()

        for i in range(2, len(vals) - 1):
            v = vals[i]
            if v == 'PRÉ':
                cha = vals[i + 1]
                if cha and str(cha).strip() not in ('X', '', 'None'):
                    result[name][str(cha).strip()] += 1
            elif v == 'CFA':
                result[name]['CFA'] += 1

    chantiers_normaux = sorted({c for emp in result.values() for c in emp if c != 'CFA'})
    all_chantiers     = chantiers_normaux + ['CFA']
    all_noms          = sorted(result.keys())
    total_col         = len(all_chantiers) + 2

    log(f"{len(all_noms)} salarié(s), {len(all_chantiers)} chantier(s) (dont CFA)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Récap chantiers"

    HDR_FILL     = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT     = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    CFA_HDR_FILL = PatternFill("solid", fgColor="7030A0")
    CFA_FILL     = PatternFill("solid", fgColor="E2CFED")
    CFA_FONT     = Font(bold=True, name="Arial", size=10, color="7030A0")
    TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
    TOTAL_FONT   = Font(bold=True, name="Arial", size=10)
    TOTAL_R_FILL = PatternFill("solid", fgColor="BDD7EE")
    GRAND_FILL   = PatternFill("solid", fgColor="9DC3E6")
    NORM_FONT    = Font(name="Arial", size=10)
    thin         = Side(style='thin', color="AAAAAA")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER       = Alignment(horizontal='center', vertical='center')
    LEFT         = Alignment(horizontal='left',   vertical='center')

    def header_cell(cell, value, fill=None):
        cell.value     = value
        cell.font      = HDR_FONT
        cell.fill      = fill or HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    last_col = openpyxl.utils.get_column_letter(total_col)
    ws.merge_cells(f'A1:{last_col}1')
    t = ws['A1']
    t.value     = f"Récap jours travaillés par chantier — {mois_label}"
    t.font      = Font(bold=True, name="Arial", size=13, color="1F4E79")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    header_cell(ws.cell(2, 1), "Salarié")
    for j, cha in enumerate(all_chantiers, 2):
        fill = CFA_HDR_FILL if cha == 'CFA' else HDR_FILL
        header_cell(ws.cell(2, j), cha, fill=fill)
    header_cell(ws.cell(2, total_col), "TOTAL")
    ws.row_dimensions[2].height = 22

    for i, nom in enumerate(all_noms, 3):
        nc = ws.cell(i, 1, nom)
        nc.font      = NORM_FONT
        nc.alignment = LEFT
        nc.border    = BORDER

        row_total = 0
        for j, cha in enumerate(all_chantiers, 2):
            val = result[nom].get(cha, 0)
            row_total += val
            cell = ws.cell(i, j, val if val > 0 else "")
            cell.alignment = CENTER
            cell.border    = BORDER

            if cha == 'CFA':
                cell.font = CFA_FONT if val > 0 else NORM_FONT
                if val > 0:
                    cell.fill = CFA_FILL
            else:
                cell.font = NORM_FONT
                if val > 0:
                    k = min(val / 22, 1.0)
                    r = int(214 - k * 100)
                    g = int(228 - k * 100)
                    b = int(255 - k * 50)
                    cell.fill = PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")

        tc = ws.cell(i, total_col, row_total)
        tc.font      = TOTAL_FONT
        tc.fill      = TOTAL_FILL
        tc.alignment = CENTER
        tc.border    = BORDER
        ws.row_dimensions[i].height = 18

    total_row = len(all_noms) + 3
    ws.cell(total_row, 1, "TOTAL").font      = TOTAL_FONT
    ws.cell(total_row, 1).fill      = TOTAL_R_FILL
    ws.cell(total_row, 1).alignment = CENTER
    ws.cell(total_row, 1).border    = BORDER

    grand_total = 0
    for j, cha in enumerate(all_chantiers, 2):
        col_total = sum(result[nom].get(cha, 0) for nom in all_noms)
        grand_total += col_total
        c = ws.cell(total_row, j, col_total)
        c.font      = TOTAL_FONT
        c.fill      = TOTAL_R_FILL
        c.alignment = CENTER
        c.border    = BORDER

    gt = ws.cell(total_row, total_col, grand_total)
    gt.font      = TOTAL_FONT
    gt.fill      = GRAND_FILL
    gt.alignment = CENTER
    gt.border    = BORDER
    ws.row_dimensions[total_row].height = 20

    ws.column_dimensions['A'].width = 26
    for j in range(2, total_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 11
    ws.freeze_panes = 'B3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
