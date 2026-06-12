"""
core/saisie_situation.py
-------------------------
Extraction des factures de situation (clients) depuis un PDF mensuel
SADAKA, et génération directe d'un fichier Excel "SUIVI DU CA" sans modèle externe.

Logique métier identique adaptée pour être appelée depuis Streamlit avec entrées/sorties en mémoire.
"""

import io
import json
import base64
import re
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

SHEET_NAME = "SUIVI DU CA"

# La ligne 5 contient désormais les en-têtes détaillés, la première ligne de données est la 6
FIRST_DATA_ROW = 6

# Mapping colonnes Excel -> clés JSON (Basé sur la structure de votre PJ)
# Colonne 1: Client, 2: Affaire, 3: Date facture, 4: N° Situation/commande, 5: n°fact, etc.
COLUMN_MAPPING = {
    1: "client",
    2: "affaire",
    3: "date_facture",
    4: "num_situation",
    5: "num_facture",
    6: "montant_ht",
    7: "tva_montant",
    8: "montant_ttc",
    9: "rg_ht",
    13: "prorata_ht",
    11: "penalites", # Ajusté selon les en-têtes standard
}

SYSTEM_PROMPT = """
Tu es un expert en comptabilité BTP française. Tu analyses des factures de
situation de travaux (PDF) émises par l'entreprise SADAKA.

Pour CHAQUE facture présente dans le document, extrait les informations
suivantes et retourne UNIQUEMENT un tableau JSON valide (aucun texte avant
ou après) avec la structure suivante :

[
  {
    "client": "nom du destinataire (donneur d'ordre)",
    "affaire": "nom du chantier / affaire",
    "date_facture": "DD/MM/YYYY",
    "num_situation": "numéro de situation (chaîne, ex: '13', '2', 'N/A')",
    "num_facture": "numéro de facture (ex: '2026/34')",
    "montant_ht": 0.0,
    "tva_montant": 0.0,
    "montant_ttc": 0.0,
    "rg_ht": 0.0,
    "penalites": 0.0,
    "prorata_ht": 0.0
  }
]

Règles importantes :
- "montant_ht" et "tva_montant" correspondent à la situation HT du mois (colonne
  "Situation HT du mois" / dernière colonne du tableau), pas au total cumulé
  ni au montant du marché.
- "montant_ttc" correspond au net à payer / montant de la situation du mois
  (dernière colonne "MONTANT TTC NET A PAYER" ou équivalent), pas au cumulé.
- "rg_ht" = retenue de garantie du mois (situation du mois), 0 si absente.
- "prorata_ht" = part de prorata HT/TTC du mois, 0 si absente.
- "penalites" = pénalités du mois, 0 si absentes.
- Pour les factures sans tableau de situation (factures simples), utilise le
  montant HT, la TVA et le montant TTC de la facture ; rg_ht, penalites et
  prorata_ht = 0.
- Si une valeur n'est pas applicable, mets 0 pour les nombres et "N/A" pour
  num_situation si non précisé.
"""

USER_PROMPT = "Extrais toutes les factures présentes dans ce document en JSON."


def extraire_donnees_pdf(pdf_path: str, api_key: str) -> list:
    """Envoie le PDF à Claude et retourne la liste des factures extraites."""
    client = anthropic.Anthropic(api_key=api_key)

    with open(pdf_path, "rb") as f:
        pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64,
                        },
                    },
                    {"type": "text", "text": USER_PROMPT},
                ],
            }
        ],
    )

    contenu = message.content[0].text.strip()
    match = re.search(r"\[.*\]", contenu, re.DOTALL)
    json_clean = match.group(0) if match else contenu

    return json.loads(json_clean)


def initialiser_structure_excel() -> Workbook:
    """Crée un Workbook openpyxl avec la structure exacte du fichier SUIVI fourni en PJ."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    
    # Styles
    font_main_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_sub_header = Font(name="Calibri", size=10, italic=True)
    font_col_header = Font(name="Calibri", size=10, bold=True)
    
    fill_blue = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_light_gray = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border_side = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Ligne 1 : En-têtes généraux de rapprochement
    ws.cell(row=1, column=1, value="Facturation mensuelle").font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=1, column=14, value="Encaissement").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=1, column=16, value="Solde créance TTC").font = Font(name="Calibri", size=11, bold=True)

    # Ligne 3 : Sous-catégories (ex: Dont)
    ws.cell(row=3, column=7, value="Dont").font = font_sub_header
    ws.cell(row=3, column=7).alignment = align_center

    # Ligne 5 : En-têtes de colonnes officiels
    headers = [
        "Client", "Affaire", "Date facture", "N° Situation / commande", "n°fact", 
        "HT", "Dont TVA", "TTC", "RG HT", "TVA RG", "Penalites", "TVA Prorata", 
        "Prorata HT", "Date Valeur banque", "Mode Rglmt", "STD / DGD", "RG", "TTC", "EG / PROMO"
    ]
    
    ws.row_dimensions[5].height = 28
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_main_header if col_idx <= 13 else font_col_header
        cell.fill = fill_blue if col_idx <= 13 else fill_light_gray
        cell.alignment = align_center
        cell.border = border_all

    return wb


def generer_excel_sans_template(factures: list) -> bytes:
    """Génère le contenu du fichier Excel directement en mémoire."""
    wb = initialiser_structure_excel()
    ws = wb[SHEET_NAME]

    thin_border_side = Side(style="thin", color="E0E0E0")
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Formats Excel de cellule
    format_currency = '#,##0.00 €'
    format_date = 'YYYY-MM-DD'

    for i, facture in enumerate(factures):
        row_idx = FIRST_DATA_ROW + i
        ws.row_dimensions[row_idx].height = 20

        for col_idx in range(1, 20):  # Initialisation des bordures de la ligne de données
            ws.cell(row=row_idx, column=col_idx).border = border_data

        for col, key in COLUMN_MAPPING.items():
            cell = ws.cell(row=row_idx, column=col)
            value = facture.get(key)

            if col in (1, 2, 4, 5):  # Textes
                cell.value = "" if value is None else str(value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 3:  # Dates
                cell.value = value if value else None
                cell.number_format = format_date
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:  # Données financières (Floats)
                try:
                    cell.value = float(value) if value not in (None, "") else 0.0
                except (TypeError, ValueError):
                    cell.value = 0.0
                cell.number_format = format_currency
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Ajustement automatique de la largeur des colonnes pour la lisibilité
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= 5 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def traiter(pdf_path: str, api_key: str, log=print) -> bytes:
    """
    Point d'entrée principal modifié (ne prend plus de template_path).
    Retourne le contenu xlsx complet en bytes.
    """
    log("Lecture du PDF par Claude Haiku...")
    factures = extraire_donnees_pdf(pdf_path, api_key)
    log(f"{len(factures)} facture(s) extraite(s).")

    log("Génération du fichier Excel de Suivi...")
    return generer_excel_sans_template(factures)
