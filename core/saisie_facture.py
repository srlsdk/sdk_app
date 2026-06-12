"""
core/saisie_facture.py
----------------------
Extraction de factures fournisseurs (BL) depuis un ou plusieurs PDF,
avec OCR de secours, et génération d'un fichier Excel récapitulatif.

Logique métier identique au script CLI d'origine (saisie_facture.py v3),
adaptée pour être appelée depuis Streamlit (entrées/sorties en mémoire).
"""

import io
import json
import re
import time
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed

import pdfplumber
import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- dépendances OCR (optionnelles) --------------------------------------------
OCR_AVAILABLE = False
try:
    from pdf2image import convert_from_path
    import pytesseract
    from PIL import ImageFilter, ImageOps
    OCR_AVAILABLE = True
except ImportError:
    pass

# --- constantes ------------------------------------------------------------------
MODEL      = "claude-haiku-4-5-20251001"
MAX_TOKENS = 8192
RETRY_WAIT = 5

OCR_MIN_CHARS_PER_PAGE   = 30
MIN_CHARS_PER_PAGE_NATIVE = 300
MAX_TEXT_CHARS           = 400_000
PAGES_PER_CHUNK          = 15

# nb de pages OCR traitées en parallèle (ajusté selon l'environnement)
OCR_WORKERS = 4

COLUMNS = [
    "Fournisseur",
    "Chantiers",
    "Bon de Livraison",
    "Date BL",
    "N° de Facture",
    "Date Facture",
    "Échéance",
    "Total H.T",
    "Total T.T.C",
]

SYSTEM_PROMPT = """Tu es un expert en extraction de données de factures françaises (secteur BTP / plomberie / chauffage).

Depuis le texte brut d'un PDF de factures fournisseurs, extrais les informations suivantes.
Retourne UNIQUEMENT un tableau JSON valide (sans markdown, sans texte avant ou après).

Chaque élément représente UN bon de livraison (BL).
Si une facture regroupe plusieurs BLs (lignes "TOTAL H.T" ou "Total du BL" distincts) → UNE ligne par BL.
Si une facture n'a qu'un seul BL → une seule ligne.

Champs :
{
  "fournisseur"  : "Nom commercial du fournisseur",
  "chantier"     : "Nom officiel complet de la commune en majuscules",
  "bon_livraison": "Numéro du BL (ou numéro de commande si pas de BL explicite)",
  "date_bl"      : "Date du BL au format DD/MM/YYYY",
  "num_facture"  : "Numéro de facture",
  "date_facture" : "Date de la facture au format DD/MM/YYYY",
  "echeance"     : "Date d'échéance DD/MM/YYYY — cherche sur TOUTES les pages",
  "total_ht"     : total HT du BL en nombre décimal (point comme séparateur décimal),
  "total_ttc"    : total TTC du BL en nombre décimal
}

RÈGLES CRITIQUES POUR LE CHAMP 'CHANTIER' :
- Déduis le chantier depuis l'adresse de livraison (code postal + ville).
- Tu dois TOUJOURS retourner le NOM OFFICIEL COMPLET de la commune en MAJUSCULES.
- Exemples de normalisation obligatoire :
  "93390 CLICHY" ou "CLICHY" -> "CLICHY-SOUS-BOIS"
  "93320 PAVILLONS" ou "PAVILLONS" -> "LES PAVILLONS-SOUS-BOIS"
  "92270 BOIS COLOMBES" -> "BOIS-COLOMBES"
  "95290 L'ISLE-ADAM" -> "L'ISLE-ADAM"
  "95610 ERAGNY" -> "ERAGNY"
  "92150 SURESNES" -> "SURESNES"
  "93420 VILLEPINTE" -> "VILLEPINTE"
  "93700 DRANCY" -> "DRANCY"
  "92170 VANVES" -> "VANVES"
  "92220 BAGNEUX" -> "BAGNEUX"
  "95800 CERGY" -> "CERGY"

AUTRES RÈGLES :
- Pour H-TUBE (factures multi-pages) : chaque BL a sa ligne "TOTAL H.T xxx".
  Le TTC = HT × 1,20. L'échéance en bas de la dernière page s'applique à TOUS les BLs.
  Si un BL s'étale sur 2 pages, attribue le montant à l'adresse de la 1re page.
- Si le même numéro de BL apparaît plusieurs fois, ne crée qu'UNE seule ligne.
- Si une valeur est introuvable → null.
- Si le total HT est 0,00 → null.
- Retourne UNIQUEMENT le tableau JSON, rien d'autre.
"""


# --- extraction texte native (pdfplumber) -----------------------------------------
def extract_pdf_text_native(pdf_path: Path) -> tuple:
    pages_text = []
    chars_par_page = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        nb_pages = len(pdf.pages)
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            pages_text.append(f"=== PAGE {i} ===\n{text}")
            chars_par_page.append(len(text.strip()))

    full = "\n".join(pages_text)
    moyenne = sum(chars_par_page) / nb_pages if nb_pages else 0
    pages_vides = sum(1 for c in chars_par_page if c < 50)

    if moyenne < MIN_CHARS_PER_PAGE_NATIVE or pages_vides > nb_pages * 0.5:
        return "", nb_pages

    return full, nb_pages


# --- OCR ----------------------------------------------------------------------------
def _ocr_one_page(args: tuple) -> tuple:
    """Fonction top-level nécessaire pour ProcessPoolExecutor."""
    page_idx, img_bytes, tesseract_config = args

    import pytesseract as _pt
    from PIL import Image, ImageFilter, ImageOps
    import io as _io

    img = Image.open(_io.BytesIO(img_bytes))
    img = img.convert("L")
    img = ImageOps.autocontrast(img, cutoff=2)
    img = img.filter(ImageFilter.SHARPEN)

    text = _pt.image_to_string(img, lang="fra", config=tesseract_config)
    return page_idx, text, len(text.strip())


def ocr_pdf(pdf_path: Path, log) -> tuple:
    """Convertit toutes les pages en images puis OCR en parallèle."""
    tesseract_config = r"--oem 1 --psm 6"

    log(f"     Conversion PDF → images (DPI 300)...")
    images = convert_from_path(str(pdf_path), dpi=300, fmt="png")
    nb_pages = len(images)
    log(f"     {nb_pages} page(s) — OCR sur {OCR_WORKERS} workers en parallèle...")

    args_list = []
    for i, img in enumerate(images, 1):
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        args_list.append((i, buf.getvalue(), tesseract_config))

    results = [None] * nb_pages
    pages_faibles = []

    with ProcessPoolExecutor(max_workers=OCR_WORKERS) as executor:
        futures = {executor.submit(_ocr_one_page, a): a[0] for a in args_list}
        for future in as_completed(futures):
            page_idx, text, chars = future.result()
            results[page_idx - 1] = (page_idx, text, chars)

    pages_text = []
    for page_idx, text, chars in results:
        if chars < OCR_MIN_CHARS_PER_PAGE:
            pages_faibles.append(page_idx)
            log(f"     [AVERTISSEMENT] Page {page_idx}/{nb_pages} : {chars} chars seulement")
        else:
            log(f"     [OK] Page {page_idx}/{nb_pages} : {chars} chars")
        pages_text.append(f"=== PAGE {page_idx} ===\n{text}")

    return "\n".join(pages_text), pages_faibles


# --- appel Claude --------------------------------------------------------------------
def _split_text_by_pages(text: str, pages_per_chunk: int) -> list:
    parts = re.split(r"(=== PAGE \d+ ===)", text)
    pages = []
    for i in range(1, len(parts), 2):
        header = parts[i]
        content = parts[i + 1] if i + 1 < len(parts) else ""
        pages.append(header + content)

    chunks = []
    for i in range(0, max(len(pages), 1), pages_per_chunk):
        chunks.append("".join(pages[i:i + pages_per_chunk]))
    return chunks if chunks else [text]


def _call_claude_once(client, text: str, filename: str, chunk_info: str, log) -> list:
    tokens_estimes = len(text) // 4
    if tokens_estimes > 40_000:
        pause = min((tokens_estimes / 50_000) * 65, 120)
        log(f"     Chunk volumineux (~{tokens_estimes:,} tokens) — pause {pause:.0f}s...")
        time.sleep(pause)

    for attempt in range(5):
        try:
            msg = client.messages.create(
                model=MODEL,
                max_tokens=MAX_TOKENS,
                system=SYSTEM_PROMPT,
                messages=[{
                    "role": "user",
                    "content": f"Fichier : {filename}{chunk_info}\n\n{text}"
                }]
            )
            raw = msg.content[0].text.strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw).strip()
            data = json.loads(raw)
            return data if isinstance(data, list) else []

        except json.JSONDecodeError as e:
            log(f"     JSON invalide (tentative {attempt+1}/5) : {e}")
            if attempt < 4:
                time.sleep(RETRY_WAIT)

        except anthropic.RateLimitError as e:
            wait = 60 * (attempt + 1)
            try:
                retry_after = int(e.response.headers.get("retry-after", wait))
                wait = max(wait, retry_after)
            except Exception:
                pass
            log(f"     Rate limit (tentative {attempt+1}/5) — attente {wait}s...")
            time.sleep(wait)

        except anthropic.APIError as e:
            log(f"     Erreur API (tentative {attempt+1}/5) : {e}")
            if attempt < 4:
                time.sleep(RETRY_WAIT * 2)

    return []


def call_claude(client, text: str, filename: str, log) -> list:
    if len(text) > MAX_TEXT_CHARS:
        log(f"     Avertissement : texte très long ({len(text)} chars), tronqué à {MAX_TEXT_CHARS}")
        text = text[:MAX_TEXT_CHARS]

    chunks = _split_text_by_pages(text, PAGES_PER_CHUNK)
    nb_chunks = len(chunks)

    if nb_chunks == 1:
        return _call_claude_once(client, chunks[0], filename, "", log)

    log(f"     {nb_chunks} tranches de ~{PAGES_PER_CHUNK} pages chacune")
    all_rows = []
    for i, chunk in enumerate(chunks, 1):
        log(f"     Tranche {i}/{nb_chunks}...")
        rows = _call_claude_once(client, chunk, filename, f" (tranche {i}/{nb_chunks})", log)
        all_rows.extend(rows)
        if i < nb_chunks:
            time.sleep(2)

    return all_rows


# --- traitement d'un PDF ---------------------------------------------------------------
def process_pdf(pdf_path: Path, api_key: str, log) -> list:
    """Retourne la liste des lignes (dict) extraites pour ce PDF."""
    log(f"\n  Fichier : {pdf_path.name}")

    text, nb_pages = extract_pdf_text_native(pdf_path)

    if text.strip():
        log(f"     Texte natif extrait ({nb_pages} pages, {len(text)} chars) → appel API")
    else:
        if not OCR_AVAILABLE:
            log("     ERREUR : OCR non disponible pour ce PDF scanné.")
            return []

        text, pages_faibles = ocr_pdf(pdf_path, log)

        if pages_faibles:
            log(f"     ERREUR OCR — Pages avec texte insuffisant : {pages_faibles}")
            log(f"     Le fichier '{pdf_path.name}' n'a pas pu être traité correctement.")
            return []

        nb_pages_ocr = text.count("=== PAGE")
        log(f"     OCR terminé : {nb_pages_ocr} pages OK, {len(text)} chars → appel API")

    client = anthropic.Anthropic(api_key=api_key)
    rows = call_claude(client, text, pdf_path.name, log)

    log(f"     → {len(rows)} ligne(s) extraite(s)")
    return rows


# --- export Excel ------------------------------------------------------------------------
def build_excel(rows: list) -> bytes:
    """Construit le fichier Excel récapitulatif et le retourne en bytes."""
    key_map = {
        "fournisseur":   "Fournisseur",
        "chantier":      "Chantiers",
        "bon_livraison": "Bon de Livraison",
        "date_bl":       "Date BL",
        "num_facture":   "N° de Facture",
        "date_facture":  "Date Facture",
        "echeance":      "Échéance",
        "total_ht":      "Total H.T",
        "total_ttc":     "Total T.T.C",
    }

    normalized = []
    for r in rows:
        norm = {}
        for src, dst in key_map.items():
            norm[dst] = r.get(src) or r.get(dst)
        normalized.append(norm)

    df = pd.DataFrame(normalized, columns=COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"

    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths = [20, 18, 22, 12, 22, 14, 12, 13, 13]

    for c_idx, (col_name, width) in enumerate(zip(COLUMNS, col_widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    alt_fill    = PatternFill("solid", start_color="EBF3FB")
    normal_fill = PatternFill("solid", start_color="FFFFFF")
    num_font    = Font(name="Calibri", size=10)
    num_align   = Alignment(horizontal="right")

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if r_idx % 2 == 0 else normal_fill
        for c_idx, value in enumerate(row, 1):
            cell        = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill   = fill
            cell.border = border
            cell.font   = num_font
            if c_idx in (8, 9):
                try:
                    cell.value         = float(value) if value is not None else None
                    cell.number_format = "#,##0.00"
                    cell.alignment     = num_align
                except (TypeError, ValueError):
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def traiter(pdf_paths: list, api_key: str, log=print) -> bytes:
    """
    Point d'entrée principal.
    pdf_paths : liste de chemins (Path) vers des PDF sur disque (temp dir).
    Retourne le contenu xlsx en bytes.
    """
    all_rows = []
    for pdf_path in pdf_paths:
        rows = process_pdf(Path(pdf_path), api_key, log)
        all_rows.extend(rows)

    return build_excel(all_rows)
